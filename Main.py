import os
from time import time

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import Parameters
from SudokuGame import SudokuGame

if __name__ == '__main__':

    stored_results = pd.DataFrame()
    stored_time = pd.DataFrame()

    for i in range(Parameters.runs):
        found = 0
        start = time()

        s = SudokuGame(Parameters.population_size, Parameters.generations, Parameters.mutation_rate)
        s.load(Parameters.currentSudoku)
        solution, resultsdf = s.play()

        end = time()
        if solution:
            if Parameters.verbose:
                s.print_solution(solution)
            found = 1
        else:
            if Parameters.verbose:
                print("NO SOLUTION WAS FOUND")

        timedf = pd.DataFrame([[Parameters.sudokuName, np.round((end - start), 6), found]])
        resultsdf = pd.DataFrame(resultsdf)

        stored_time = pd.concat((stored_time, timedf))
        stored_results = pd.concat((stored_results, resultsdf))

    stored_time.to_excel(Parameters.timePath, index=False)
    stored_results.to_excel(Parameters.resultsPath, index=False)

'''
    time_path = open("time_analysis.xlsx", "r")
    results_path = open("results.xlsx", "r")

    with pd.ExcelWriter(time_path,engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
        writer.book = load_workbook(time_path)
        stored_time.to_excel(writer, sheet_name="Sheet1", startrow=writer.sheets['Sheet1'].max_row, header=None)

    with pd.ExcelWriter(results_path, engine="openpyxl", mode='a', if_sheet_exists="overlay") as writer:
        writer.book = load_workbook(results_path)
        stored_results.to_excel(writer, sheet_name="Sheet1", startrow=writer.sheets['Sheet1'].max_row, header=None)
'''

